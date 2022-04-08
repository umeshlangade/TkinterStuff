from tkinter import *
from openpyxl import Workbook

root = Tk()

def callback():
    label.config(text="You clicked me, Umesh!!", fg="white", bg="black")
    wb = Workbook()
    ws = wb.active
    cellx='A1'
    print(ws[cellx].value)
    ws[cellx]='UMESH SHANTIR LANGADE'
    print(ws[cellx].value)
    wb.save("sample.xlsx")

#LABEL
label = Label(root,text="Hello Tkinter")
label.pack()

#BUTTON
button = Button(root,text="Click me..",command=callback)
#button['state']='disabled'
#button['state']='normal'
button.pack()

root.geometry("350x300")
root.mainloop()